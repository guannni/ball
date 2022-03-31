#  给出旋转的traj；顺势轴和角度，euler，四元数，angular displacement；主要用于计算PDF（方法一）
#   'D:\\guan2019\\2_ball\\2_data\\60Hz_select_rot\\' 是用来画rotational pdf的！！！！！

import tables as tb
import math
import pandas as pd
import trackpy as tp
import matplotlib.pyplot as plt
import numpy as np
import os.path
import warnings
import sympy as sp
import matplotlib
import matplotlib.cm as cm
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.ticker  import MultipleLocator
from matplotlib.ticker import FuncFormatter
import openpyxl


warnings.filterwarnings('ignore')
# TODO: CHANGE PARAMETERS HERE------------------
fps = 150.0


# -----------------------------------------------
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)



def compute_msd(trajectory, t_step, coords=['x', 'y']):
    tau = trajectory['t'].copy()
    shifts = np.floor(tau / t_step).astype(np.int)
    msds = np.zeros(shifts.size)
    msds_std = np.zeros(shifts.size)

    for i, shift in enumerate(shifts):
        diffs = trajectory[coords] - trajectory[coords].shift(-shift)
        sqdist = np.square(diffs).sum(axis=1)
        msds[i] = sqdist.mean()
        msds_std[i] = sqdist.std()

    msds = pd.DataFrame({'msds': msds, 'tau': tau, 'msds_std': msds_std})
    return msds


def traversalDir_FirstDir(path):  # 返回一级子文件夹名字
    list = []
    if (os.path.exists(path)):  # 获取该目录下的所有文件或文件夹目录路径
        files = glob.glob(path + '\\*')
        # print(files)
        for file in files:  # 判断该路径下是否是文件夹
            if (os.path.isdir(file)):  # 分成路径和文件的二元元组
                h = os.path.split(file)
                print(h[1])
                list.append(h[1])
        return list



# read position from hdf5 file
path2 = 'D:\\guan2019\\2_ball\\2_data_new\\60Hz_copy\\'  # _rot_selected\\' # select_rot\\'  # TODO: 注意 这里只用_copy文件夹的数据！！！！

# 存每组数据pdf的文件 （全篇搜索path4，还有一段同时注释掉的）
# path4 = 'D:\\guan2021\\report_b\\temp.xlsx'
# write_excel_xlsx(path4,'temp',tuple([np.arange(-30.5, 30, 1)])) 

err_rot_dict = {}
err_rot_dict['3.5']=[0,	0,	0,	0,	0,	0	,5.57374E-08,	5.33231E-08,	5.05262E-08,	5.57374E-08	,1.4854E-08,	1.97201E-06,	1.4961E-06,	1.15184E-06	,3.39518E-06,	4.24149E-06,	2.25886E-06,	6.76048E-06	,0.0000278063,	0.0000453364	,0.000065724,	0.00009376051,	0.0002039575,	0.0009232703	,0.0017276	,0.001429606,	0.0030780055,	0.009041447,	0.009089505,		0.004269722,0.008089505,0.009089505,	0.008008378,	0.008041108	,0.0090201062,	0.002141019,	0.00175018787,	0.0005348621,	0.0005100042,	0.00008406888,	0.0000583191,	0.0000514562	,0.00000292859	,6.82248E-07	,2.46004E-7,	4.02222E-07	,6.15611E-07,	1.05734E-07	,1.4961E-07,	5.33231E-08,	9.65403E-08,	5.59863E-08,	5.33231E-08,	5.05262E-08,	0,	0,0,	0,5.57374E-08,	0]
err_rot_dict['4.0']=[0,	0,	0,	0,	5.57374E-06,	0	,5.57374E-06,	5.33231E-06,	5.05262E-06,	5.57374E-06	,1.4854E-05,	1.97201E-05,	1.4961E-05,	1.15184E-05	,3.39518E-05,	4.24149E-05,	2.25886E-05,	6.76048E-05	,0.000278063,	0.000453364	,0.00065724,	0.001376051,	0.002039575,	0.003232703	,0.005076	,0.008929606,	0.010780055,	0.009041447,	0.016089505,	0.004269722	,0.009483024,	0.003868066,	0.015008378,	0.008041108	,0.010201062,	0.008141019,	0.005018787,	0.003348621,	0.002100042,	0.001406888,	0.000683191,	0.000514562	,0.000292859	,6.82248E-05	,2.46004E-05,	4.02222E-05	,6.15611E-05,	1.05734E-05	,1.4961E-05,	5.33231E-06,	9.65403E-06,	5.59863E-06,	5.33231E-06,	5.05262E-06,	0,	0,	0,	0,5.57374E-06,	0]
err_rot_dict['4.5']=[0,	6.46872E-06,	0,	1.31135E-05	,4.50958E-06	,9.01917E-06	,0	,0,	9.80068E-06,	6.89892E-06	,1.39346E-05,	1.45871E-05	,1.49016E-05	,1.40105E-05	,2.1979E-05,	1.91528E-05	,3.71155E-05	,6.197E-05	,8.47205E-05	,0.000226265,	0.000390342	,0.000692235	,0.000947072	,0.001859938	,0.00293451,	0.006944663,	0.006157939	,0.003610306,	0.010981695,	0.003919687,	0.011408825,	0.005461761,	0.010573001,0.00450112,	0.005759442,	0.00650736,	0.003648957,	0.001485706,	0.001019843	,0.00064016,	0.000384049,	0.0002181	,0.00014965	,6.35352E-05,	1.55228E-05	,2.79788E-05	,1.63848E-05,	1.5227E-05,	1.7312E-05,	2.39637E-05	,1.31868E-05	,9.80068E-06	,4.50958E-06,	0	,0	,0,	1.31311E-05	,0,	0,	0]
err_rot_dict['5.0']=[0,	0,	5.20579E-06,	0,	0,	0	,2.08232E-05	,0,	1.04116E-05	,9.29911E-06	,9.63364E-06,	1.15905E-05,	2.1374E-05,	1.82061E-05,	4.33206E-05	,3.63321E-05,	5.2078E-05,	8.46062E-05,	9.25997E-05	,0.000264727,	0.000551671,	0.00107013,0.001600761,	0.003027858	,0.006153843,	0.008600961,	0.008177046	,0.003545573,	0.015295178	,0.010131611,	0.026154795,	0.01178008,	0.016946846,	0.003313826,	0.008538414,0.007293828	,0.005993512	,0.002883311,	0.001526209	,0.001257751,	0.000468391	,0.000248619	,0.000102925	,0.000123137	,5.89174E-05,	3.94937E-05,	3.06457E-05,	1.67626E-05,	1.04116E-05,	1.15905E-05,	1.04116E-05,	1.53756E-05	,0	,0,	0,	1.04116E-05	,9.63364E-06	,5.20579E-06,	0	,9.63364E-06]
err_rot_dict['5.5']=[0,	0,	5.15952E-06,	0,	7.53485E-06,	0,	5.56579E-06,	0	,2.89204E-06	,0	,6.43058E-06	,1.05945E-05,	5.39981E-06,	1.5933E-05,	1.95021E-05	,1.53595E-05,	1.26485E-05,	4.97115E-05	,3.8826E-05,	6.93546E-05	,0.000134671	,0.000257364	,0.000456776,	0.001104552,	0.002245641,	0.004954609,	0.002269824	,0.007665327,	0.007642214,	0.006177446,	0.018616181	,0.005817974	,0.006620894	,0.007670426,	0.003019583	,0.004630517,	0.002401807,	0.001168494	,0.000529424,	0.000272127	,0.000251335	,7.56688E-05,	6.23201E-05	,3.69235E-05,	1.65471E-05	,2.63698E-05,	2.55672E-05	,1.51263E-05	,9.19784E-06	,6.20114E-06	,5.09112E-06	,7.58137E-06	,5.56579E-06,	5.15952E-06,	8.1437E-06,	5.15952E-06,	0,	0,	0,	0]
err_rot_dict['6.0']=[4.21514E-06,	4.21514E-06,	0	,4.21514E-06,	0,	0	,2.9034E-05,	2.89891E-05	,2.89891E-05,	2.95517E-05,	0.000116151	,3.18331E-05	,2.07057E-05,	2.72647E-05	,3.75403E-05,	3.05726E-05,	0.00013483	,0.000144712,	0.000150035,	9.33529E-05,	0.000157056	,0.000243537,	0.000595254	,0.000686761,	0.002059601	,0.002354132,	0.005894316,	0.008404389,	0.004321167	,0.005241115,	0.013767557	,0.004393366	,0.005617708	,0.006663887,	0.00494132,	0.00188967	,0.002117426,	0.001299324,	0.000596815,	0.000189253	,0.000162039	,0.000100284	,5.68769E-05,	0.000106387	,4.0597E-05,	2.86704E-05,	4.31073E-05,	6.01763E-05	,1.26454E-05	,8.34768E-05,	1.26454E-05	,5.91035E-05,	0	,2.9034E-05,	2.96834E-05,	8.43028E-06,	0,	0,	0	,1.28029E-05]

err_rotenergy_dict = {}
for key in err_rot_dict:
    err_rotenergy_dict[key] = [0.5*9e-7*(a/180*math.pi)**2 for a in err_rot_dict[key]]  # J in unit


filename = [name for name in os.listdir(path2)]
pdf_rot_dict = {}

step = 1  #30

for j in range(len(filename)):  #3,4):    # len(filename)-1,
    path3 = path2 + filename[j] + '\\'
    filename1 = [os.path.splitext(name)[0] for name in os.listdir(path3)]
    file_n = [path3 + name + '.h5' for name in filename1]
    print(filename1, file_n)

    d_theta = []
    d_pr = []
    d_axis = [[0,0,0]]
    d_euler = [[0,0,0]]
    d_quaternions = [[0,0,0,0]]


    for i in range(len(file_n)):
        store = pd.HDFStore(file_n[i], mode='r')
        print(store.keys())
        center = store.get('center').values  # numpy array
        matrix = store.get('matrix').values
        points = store.get('points').values[::step]  # timestep 1/5s
        store.close()

        N = len(center)
        max_time = N / fps  # seconds
        frame_name = filename[i].split('_', 1)[0]  # 频率 为.h5文件的key，后面多组数据作图用key来挑选！！！
        print(type(frame_name))

    # # translational-----------------------------------------------------------------------------------------

        # # traj pic
        # traj = pd.DataFrame({'t': np.linspace(0, max_time, N), 'x': center[:, 0], 'y': center[:, 1]})
        # print(traj.head())
        # ax = traj.plot(x='x', y='y', alpha=0.6, legend=False, title='trajectory')
        # ax.set_xlim(0, 480)  # (traj['x'].min()-10, traj['x'].max()+10)
        # ax.set_ylim(0, 480)  # (traj['y'].min()-10, traj['y'].max()+10)
        # ax.set_xlabel('x(pixel)')
        # ax.set_ylabel('y(pixel)')
        # ax.plot()
        # plt.show()
        # fig = ax.get_figure()
        # # fig.savefig(frame_traj[i])
        # del ax, fig




    # rotational--------------------------------------------------------------------------------------------------------------
        points_reshape = np.reshape(points, (len(points), 6, 3))  # points 2维，points_reshape 3维

        if step != 1:
            # 重新算更改步长后的matrix-------------
            matrix = []
            for k in range(1, len(points)):
                points_co = points_reshape[k-1][0:3,:]
                points_ps_n = points_reshape[k][0:3,:]
                pi = sp.Matrix(points_co).T  # 转置
                pf = sp.Matrix(points_ps_n).T
                if pi.det() != 0:  # del()不为0
                    rot = pf * (pi.inv())  # 旋转矩阵
                    matrix.append(rot)


        matrix_reshape = np.reshape(np.array(matrix), (len(np.array(matrix)), 3, 3))  # reshape的矩阵
        points_1 = points_reshape[0][0]
        print(points_1)
        for k in range(1,len(points)-1):
            # print(np.linalg.norm(points_reshape[k][0]-points_1))
            # print(points_1)
            if np.linalg.norm(points_reshape[k][0] - points_1) > 10:
                # print(points_reshape[k][0],k)
                points_reshape[k] = points_reshape[k - 1]
                matrix_reshape[k - 1] = [[1, 0, 0], [0, 1, 0], [0, 0, 1]]
                matrix_reshape[k] = [[1, 0, 0], [0, 1, 0], [0, 0, 1]]

            points_1 = points_reshape[k][0]


        # # ---- 计算角度
        # # -------- 瞬时角+瞬时轴+欧拉角+四元数 --计算
        print(np.arccos(float(np.ndarray.trace(matrix_reshape[0]) - 1) / 2.))
        points_theta = np.nan_to_num(np.array([np.arccos(float(np.ndarray.trace(x) - 1) / 2.)+0.00001 for x in
                                 matrix_reshape]) ) # points_theta = arccos((tr(matrix)-1)/2) 弧度
        points_theta1 = np.array([[matrix_reshape[x][2, 1] - matrix_reshape[x][1, 2],
                                   matrix_reshape[x][0, 2] - matrix_reshape[x][2, 0],
                                   matrix_reshape[x][1, 0] - matrix_reshape[x][0, 1]] for x in range(len(matrix_reshape))])  # 瞬时角 弧度

        points_axis = np.array([points_theta1[x] / (2 * math.sin(points_theta[x])) for x in range(
            len(matrix_reshape))])  # 瞬时轴 axis = [R[2,1]-R[1,2],R[0,2]-R[2,0],R[1,0]-R[0,1]]/(2*SIN(THETA))
        deltaeuler = np.array([[math.atan2(x[2, 1], x[2, 2]),
                                  math.atan2(-x[2, 0], math.sqrt(x[2, 1] ** 2 + x[2, 2] ** 2)),
                                  math.atan2(x[1, 0], x[0, 0])] for x in matrix_reshape])  # euler 弧度 （旋转矩阵转欧拉角查公式

        quaternions = np.array([[math.sqrt(np.ndarray.trace(matrix_reshape[x])+1)/2.0,
                                 (matrix_reshape[x][1, 2] - matrix_reshape[x][2, 1]) / 2.0 / math.sqrt(
                                     np.ndarray.trace(matrix_reshape[x]) + 1),
                                 (matrix_reshape[x][2, 0] - matrix_reshape[x][0, 2]) / 2.0 / math.sqrt(
                                     np.ndarray.trace(matrix_reshape[x]) + 1),
                                 (matrix_reshape[x][0, 1] - matrix_reshape[x][1, 0]) / 2.0 / math.sqrt(
                                     np.ndarray.trace(matrix_reshape[x]) + 1)] for x in range(len(matrix_reshape))])

        # ------------瞬时角+瞬时轴+欧拉角 +四元数--处理
        deltatheta = points_theta / math.pi * 180.0
        deltaeuler = deltaeuler / math.pi * 180.0  # euler角度


        all = np.vstack((deltatheta, points_axis[:, 0], points_axis[:, 1],
                         points_axis[:, 2], deltaeuler[:, 0], deltaeuler[:, 1], deltaeuler[:, 2], quaternions[:, 0],
                         quaternions[:, 1], quaternions[:, 2], quaternions[:, 3]))  # theta[0,:],axis[1:4,:],euler[4:7,:],quaterions[7:11,:]


        index = np.where(deltatheta == 0.0)
        all = np.delete(all, index[0],axis=1)
        deltaaxis = np.delete(points_axis, index[0],axis=0)  # 瞬时轴
        index = np.where(np.nan_to_num(deltaaxis[:,0] == 0))
        all = np.delete(all, index[0],axis=1)

        index = []
        for k in range(1,len(all[0,:])):
            if (all[:,k]==all[:,k-1]).all():
                index.append(k)

        all = np.delete(all, index,axis=1)

        deltatheta = all[0,:].T  # 每次旋转的角度 (degree)
        deltaaxis = all[1:4,:].T  # 瞬时轴
        deltaeuler = all[4:7,:].T # euler (degree)
        quaternions = all[7:11,:].T  # quaternions

        # # -------- 单个点位移 --计算
        deltar = np.diff(points_reshape[:, 0, :], axis=0)  #------------单个点位移
        deltapr = np.sqrt(np.sum(deltar ** 2, axis=1))  # 每次旋转的球面距离(pixel)
        index = np.where(deltapr == 0)
        deltapr = np.delete(deltapr, index[0])

        d_theta = np.hstack((d_theta, deltatheta)) #全正theta
        d_axis = np.vstack((d_axis, deltaaxis))[1:,:]  # 去掉一开始那个[0,0,0]
        d_euler = np.vstack((d_euler, deltaeuler))[1:,:]  # 去掉一开始那个[0,0,0]
        d_pr = np.hstack((d_pr, deltapr))

        index = np.where(d_axis[:,2] <= 0)  # 正负theta（轴z+）
        d_new_theta = d_theta.copy()
        d_new_theta[index] = d_new_theta[index]*(-1)  # 把瞬时轴指向z=0以下的旋转角设为负
        d_new_axis = d_axis.copy()  # d_axis旋转轴指向z>0/<0都有，d_theta只有顺时针
        d_new_axis[index] = d_new_axis[index]*(-1)  # d_new_axis旋转轴全部指向z>0，d_new_theta有顺/逆时针

        d_quaternions = np.vstack((d_quaternions,quaternions)) # 四元数

        # # --计算本组数据正负theta（轴z+）的pdf,来计算errorbar，存到path4-----------------------------------------------------
        # d_axis1= deltaaxis
        # index1 = np.where(d_axis1[:,2] <= 0)
        # d_new_theta1 = deltatheta.copy()  #本组数据的d_new_theta
        # d_new_theta1[index1] = d_new_theta1[index1]*(-1)  # 把瞬时轴指向z=0以下的旋转角设为负
        # d_new_axis1 = d_axis1.copy()  # d_axis旋转轴指向z>0/<0都有，d_theta只有顺时针
        # d_new_axis1[index1] = d_new_axis1[index1]*(-1)  # d_new_axis旋转轴全部指向z>0，d_new_theta有顺/逆时针
        # weights_new_theta1 = np.ones_like(d_new_theta1) / float(len(d_new_theta1))
        # cuts_m_new_theta1 = int(max(-min(d_new_theta1), max(d_new_theta1)) / 0.5 // 2) + 0.5
        # cuts_new_theta1 = np.arange(-30.5, 30, 1)  # 轴指向z+的theta # 另一种bin np.arange(-30, 30, 1)
        # p_new_theta1, x_new_theta1, cu1 = plt.hist(d_new_theta1, cuts_new_theta1, histtype='bar', facecolor='yellowgreen', weights=weights_new_theta1, alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
        # x_new_theta1 = (x_new_theta1[:-1] + x_new_theta1[1:]) / 2.
        # #存每组pdf
        # data = openpyxl.load_workbook(path4)
        # sheetnames = data.get_sheet_names()
        # table = data.get_sheet_by_name('temp')
        # table = data.active
        # table.append(list(p_new_theta1))
        # data.save(path4)


        # # --------------------------------------------------------------------------------------------------------------------




        # ---- calculate -----------------------------------------------------------------------------------------------------

        # # # -- 作图rot_traj pic
        # points_traj = pd.DataFrame(
        #     {'t': np.linspace(0, max_time, N), 'x': points_reshape[:, 0,0], 'y': points_reshape[:, 0,1], 'z': points_reshape[:, 0,2]})
        # # 更改timestep-----------------------
        # px = points_traj['x'][::100]
        # py = points_traj['y'][::100]
        # pz = points_traj['z'][::100]
        # # Make ball
        # br = np.floor(np.sqrt(px[0] ** 2 + py[0] ** 2 + pz[0] ** 2)) - 2
        # bu = np.linspace(0, 2 * np.pi, 1000)
        # bv = np.linspace(0, 2 * np.pi, 1000)
        # bx = br * np.outer(np.cos(bu), np.sin(bv))
        # by = br * np.outer(np.sin(bu), np.sin(bv))
        # bz = br * np.outer(np.ones(np.size(bu)), np.cos(bv))
        # fig = plt.figure()
        # ax = fig.add_subplot(111, projection='3d')
        # ax.plot_surface(bx, by, bz, color='w', alpha=0.1)
        # for k in range(0,len(pz) - 1): # range(0,len(pz) - 1):
        #     im = ax.plot(px[k:k + 2], py[k:k + 2], pz[k:k + 2], color=plt.cm.jet(int(k * 255 / (len(pz)-1))),alpha=0.5)
        # plt.show()
        # fig = ax.get_figure()
        # del ax, fig

        # #  -- 作图theta/dtheta-time fluctuation
        # N = len(d_new_theta)
        # max_time = N / fps  # seconds
        # time = np.linspace(0, N-1, N) *step/150.0 # (0, max_time, N)
        # fig = plt.figure(figsize=(6,3))
        # ax0 = plt.subplot(111)
        # plt.scatter(time,d_new_theta, s=2) # Displacement Position pixel # mm / 480.0 * 260
        # ax0.set_xlabel('time (s)')
        # ax0.set_xlim(0,max(time))
        # ax0.set_ylabel('$\Delta \Theta$ (degree)')
        # ax0.grid(which='major', axis='x', linewidth=0.75, linestyle='-', color='0.95')  
        # plt.show()
        # fig = plt.figure(figsize=(6,3))
        # ax0 = plt.subplot(111)
        # plt.scatter(time,np.cumsum(d_new_theta), s=2) # Displacement Position pixel # mm / 480.0 * 260
        # ax0.set_xlabel('time (s)')
        # ax0.set_xlim(0,max(time))
        # ax0.set_ylabel('$\Theta$ (degree)')
        # ax0.grid(which='major', axis='x', linewidth=0.75, linestyle='-', color='0.95')  
        # plt.show()

        # #  -- 作图euler angles-time fluctuation
        # N = len(d_euler[:,0])
        # max_time = N / fps  # seconds
        # time = np.linspace(0, N-1, N) *step/150.0 # (0, max_time, N)
        # d_euler1 = d_euler[:,0]
        # d_euler2 = d_euler[:,1]
        # d_euler3 = d_euler[:,2]
        # fig = plt.figure(figsize=(6,3))
        # ax0 = plt.subplot(111)
        # plt.scatter(time,d_euler1, s=2) 
        # # plt.plot(time,d_euler1,linewidth=0.5) 
        # ax0.set_xlabel('time (s)')
        # ax0.set_xlim(0,max(time))
        # ax0.set_ylabel(r'$\Delta \alpha $ (degree)')
        # ax0.grid(which='major', axis='x', linewidth=0.75, linestyle='-', color='0.95')  
        # plt.show()
        # fig = plt.figure(figsize=(6,3))
        # ax0 = plt.subplot(111)
        # plt.scatter(time,d_euler2, s=2) 
        # # plt.plot(time,d_euler2) 
        # ax0.set_xlabel('time (s)')
        # ax0.set_xlim(0,max(time))
        # ax0.set_ylabel(r'$\Delta \beta $ (degree)')
        # ax0.grid(which='major', axis='x', linewidth=0.75, linestyle='-', color='0.95')  
        # plt.show()
        # fig = plt.figure(figsize=(6,3))
        # ax0 = plt.subplot(111)
        # plt.scatter(time,d_euler3, s=2) 
        # # plt.plot(time,d_euler3) 
        # ax0.set_xlabel('time (s)')
        # ax0.set_xlim(0,max(time))
        # ax0.set_ylabel(r'$\Delta \gamma $ (degree)')
        # ax0.grid(which='major', axis='x', linewidth=0.75, linestyle='-', color='0.95')  
        # plt.show()






        # # # -- msd
        # # dt = max_time/N
        # # msd = compute_msd(traj, t_step=dt, coords=['x', 'y'])
        # # print(msd.head())
        # # ax = msd.plot(x="tau", y="msds", logx=True, logy=True, legend=False, title='MSD')
        # # ax.fill_between(msd['tau'], msd['msds'] - msd['msds_std'], msd['msds'] + msd['msds_std'], alpha=0.2)
        # # ax.plot()
        # # plt.show()
        # # msd_i = pd.HDFStore(frame_msd[i], complib='blosc')
        # # msd_i.append(frame_name, msd, format='t', data_columns=True)
        # # fig = ax.get_figure()
        # # # fig.savefig(frame_msd_pic[i])
        # # del msd, msd_i, ax, fig



#   # pr
#     weights_pr = np.ones_like(d_pr) / float(len(d_pr))
#     cuts_m_pr = int(max(-min(d_pr), max(d_pr)) / 0.5 // 2) + 0.5
#     cuts_pr = np.arange(-0.025, max(d_pr), 0.01)
#     p_pr, x_pr, cu = plt.hist(d_pr, cuts_pr, histtype='bar', facecolor='yellowgreen', weights=weights_pr, alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
#     pdf_rot_dict[filename[j]+'p_pr'] = p_pr
#     x_pr = (x_pr[:-1] + x_pr[1:]) / 2.
#     pdf_rot_dict[filename[j]+'x_pr'] = x_pr  # 存入dict

# # axis-angle-----------------------------------------------------------------------
#   # 全正theta --用了两种bar来画
#     weights_theta = np.ones_like(d_theta) / float(len(d_theta))
#     cuts_m_theta = int(max(-min(d_theta), max(d_theta)) / 0.5 // 2) + 0.5
#     cuts_theta = np.arange(-0.5, 30, 1) # 轴指向z-z+都有的theta # 另一种bin  np.arange(0, 30, 1)
#     p_theta, x_theta, cu = plt.hist(d_theta, cuts_theta, histtype='bar', facecolor='yellowgreen', weights=weights_theta, alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
#     pdf_rot_dict[filename[j]+'p_theta'] = p_theta
#     x_theta = (x_theta[:-1] + x_theta[1:]) / 2.
#     pdf_rot_dict[filename[j]+'x_theta'] = x_theta  # 存入dict

  # 正负theta（轴z+）--用了两种bar来画
    weights_new_theta = np.ones_like(d_new_theta) / float(len(d_new_theta))
    cuts_m_new_theta = int(max(-min(d_new_theta), max(d_new_theta)) / 0.5 // 2) + 0.5
    cuts_new_theta = np.arange(-30.5, 30, 1)  # 轴指向z+的theta # 另一种bin np.arange(-30.5, 30, 1)
    p_new_theta, x_new_theta, cu = plt.hist(d_new_theta, cuts_new_theta, histtype='bar', facecolor='yellowgreen', weights=weights_new_theta, alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
    pdf_rot_dict[filename[j]+'p_new_theta'] = p_new_theta/sum(p_new_theta)
    x_new_theta = (x_new_theta[:-1] + x_new_theta[1:]) / 2.
    pdf_rot_dict[filename[j]+'x_new_theta'] = x_new_theta  # 存入dict

  # 正负theta（轴z+）对应的rotationalenergy
    rot_energy = 0.5*9e-7*(d_new_theta/180*math.pi)**2  # unit J
    weights_rot_energy = np.ones_like(rot_energy) / float(len(rot_energy))
    cuts_m_rot_energy = int(max(-min(rot_energy), max(rot_energy)) / 0.5 // 2) + 0.5
    cuts_rot_energy = np.arange(-1e-10, 9E-9, 3E-10)  
    p_rot_energy, x_rot_energy, cu = plt.hist(rot_energy, cuts_rot_energy, histtype='bar', density=True)  #   facecolor='yellowgreen',weights=weights_rot_energy,  alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
    pdf_rot_dict[filename[j]+'p_rot_energy'] = p_rot_energy/sum(p_rot_energy)
    x_rot_energy = (x_rot_energy[:-1] + x_rot_energy[1:]) / 2.
    pdf_rot_dict[filename[j]+'x_rot_energy'] = x_rot_energy  # 存入dict

    err_rotenergy_dict[filename[j]] = [a/sum(p_rot_energy) for a in err_rotenergy_dict[filename[j]]]  # J in unit


# euler----------------------------------------------------------------------------
  # euler1
    d_euler1 = d_euler[:,0]
    weights_euler1 = np.ones_like(d_euler1) / float(len(d_euler1))
    cuts_m_euler1 = int(max(-min(d_euler1), max(d_euler1)) / 0.5 // 2)# + 0.5
    cuts_euler1 = np.arange(-30.5, 30, 1)
    p_euler1, x_euler1, cu = plt.hist(d_euler1, cuts_euler1, histtype='bar', facecolor='yellowgreen', weights=weights_euler1, alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
    pdf_rot_dict[filename[j]+'p_euler1'] = p_euler1
    x_euler1 = (x_euler1[:-1] + x_euler1[1:]) / 2.
    pdf_rot_dict[filename[j]+'x_euler1'] = x_euler1  # 存入dict
  # euler2
    d_euler2 = d_euler[:,1]
    weights_euler2 = np.ones_like(d_euler2) / float(len(d_euler2))
    cuts_m_euler2 = int(max(-min(d_euler2), max(d_euler2)) / 0.5 // 2)# + 0.5
    cuts_euler2 = np.arange(-30.5, 30, 1)
    p_euler2, x_euler2, cu = plt.hist(d_euler2, cuts_euler2, histtype='bar', facecolor='yellowgreen', weights=weights_euler2, alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
    pdf_rot_dict[filename[j]+'p_euler2'] = p_euler2
    x_euler2 = (x_euler2[:-1] + x_euler2[1:]) / 2.
    pdf_rot_dict[filename[j]+'x_euler2'] = x_euler2  # 存入dict
  # euler3
    d_euler3 = d_euler[:,2]
    weights_euler3 = np.ones_like(d_euler3) / float(len(d_euler3))
    cuts_m_euler3 = int(max(-min(d_euler3), max(d_euler3)) / 0.5 // 2)# + 0.5
    cuts_euler3 = np.arange(-30.5, 30, 1)
    p_euler3, x_euler3, cu = plt.hist(d_euler3, cuts_euler3, histtype='bar', facecolor='yellowgreen', weights=weights_euler3, alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
    pdf_rot_dict[filename[j]+'p_euler3'] = p_euler3
    x_euler3 = (x_euler3[:-1] + x_euler3[1:]) / 2.
    pdf_rot_dict[filename[j]+'x_euler3'] = x_euler3  # 存入dict

# # quaternions-----------------------------------------------------------
#     # quaternion1
#     d_quat1 = d_quaternions[:,0]
#     weights_quat1= np.ones_like(d_quat1) / float(len(d_quat1))
#     cuts_m_quat1 = int(max(-min(d_quat1), max(d_quat1)) / 0.5 // 2)# + 0.5
#     cuts_quat1 = np.arange(-0.01,1,0.02)
#     p_quat1, x_quat1, cu = plt.hist(d_quat1, cuts_quat1, histtype='bar', facecolor='yellowgreen',weights=weights_quat1, alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
#     pdf_rot_dict[filename[j] + 'p_quat1'] = p_quat1
#     x_quat1 = (x_quat1[:-1] + x_quat1[1:]) / 2.
#     pdf_rot_dict[filename[j] + 'x_quat1'] = x_quat1  # 存入dict
#     # quaternion2
#     d_quat2 = d_quaternions[:,1]
#     weights_quat2= np.ones_like(d_quat2) / float(len(d_quat2))
#     cuts_m_quat2 = int(max(-min(d_quat2), max(d_quat2)) / 0.5 // 2)# + 0.5
#     cuts_quat2 = np.arange(-0.201,0.2,0.002)
#     p_quat2, x_quat2, cu = plt.hist(d_quat2, cuts_quat2, histtype='bar', facecolor='yellowgreen',weights=weights_quat2, alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
#     pdf_rot_dict[filename[j] + 'p_quat2'] = p_quat2
#     x_quat2 = (x_quat2[:-1] + x_quat2[1:]) / 2.
#     pdf_rot_dict[filename[j] + 'x_quat2'] = x_quat2  # 存入dict
#      # quaternion3
#     d_quat3 = d_quaternions[:,2]
#     weights_quat3= np.ones_like(d_quat3) / float(len(d_quat3))
#     cuts_m_quat3 = int(max(-min(d_quat3), max(d_quat3)) / 0.5 // 2)# + 0.5
#     cuts_quat3 = np.arange(-0.201,0.2,0.002)
#     p_quat3, x_quat3, cu = plt.hist(d_quat3, cuts_quat3, histtype='bar', facecolor='yellowgreen',weights=weights_quat3, alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
#     pdf_rot_dict[filename[j] + 'p_quat3'] = p_quat3
#     x_quat3 = (x_quat3[:-1] + x_quat3[1:]) / 2.
#     pdf_rot_dict[filename[j] + 'x_quat3'] = x_quat3  # 存入dict
#     # quaternion1
#     d_quat4 = d_quaternions[:,3]
#     weights_quat4= np.ones_like(d_quat4) / float(len(d_quat4))
#     cuts_m_quat4 = int(max(-min(d_quat4), max(d_quat4)) / 0.5 // 2)# + 0.5
#     cuts_quat4 = np.arange(-0.101,0.1,0.002)
#     p_quat4, x_quat4, cu = plt.hist(d_quat4, cuts_quat4, histtype='bar', facecolor='yellowgreen',weights=weights_quat4, alpha=0.75, rwidth=1, density=True)  # au是counts，bu是deltar
#     pdf_rot_dict[filename[j] + 'p_quat4'] = p_quat4
#     x_quat4 = (x_quat4[:-1] + x_quat4[1:]) / 2.
#     pdf_rot_dict[filename[j] + 'x_quat4'] = x_quat4  # 存入dict


# -----------pdf汇总作图-------------
ys = [i +  (i ) ** 2 for i in range(len(filename))]
colors = cm.rainbow(np.linspace(0, 1, len(ys)))
marker = ['o', 'v', 'D','^','s', 'h', '2', 'p', '*',  '+', 'x']

# # -- rot_pdf
# # 点的r的pdf
# # path2 = 'D:\\guan2019\\2_ball\\2_data\\60Hz_copy\\'
# fig = plt.figure()
# ax0 = plt.subplot(111)
# label = []
# for i in range(len(filename)):  # 2,3):    #
#     plt.plot(pdf_rot_dict[filename[i] + 'x_pr'] / 480.0 * 260, pdf_rot_dict[filename[i] + 'p_pr'], alpha=0.5,color=colors[i])
#     plt.scatter(pdf_rot_dict[filename[i] + 'x_pr'] / 480.0 * 260, pdf_rot_dict[filename[i] + 'p_pr'], marker=marker[i], c='', s=25, edgecolor=colors[i], label=filename[i] + 'g')
#     label.append(filename[i] + 'g')
# ax0.set_title('PDF -- the Value of Single Point\'s Displacement' + ' [60Hz] ', fontsize=10)
# leg0 = plt.legend(label)
# leg0.get_frame().set_linewidth(0.0)
# ax0.xaxis.set_minor_locator(MultipleLocator(0.05))
# ax0.yaxis.set_minor_locator(MultipleLocator(0.5))
# ax0.xaxis.set_major_locator(MultipleLocator(0.1))
# ax0.tick_params(axis="x", direction="in")
# ax0.tick_params(axis="y", direction="in")
# ax0.tick_params(which='minor', direction='in')
# ax0.set_xlabel('$\Delta R(mm)$')  # ax0.set_xlabel('deltaR(pixel)')
# ax0.set_ylabel('P')
# ax0.plot()
# plt.show()
# del ax0, fig

# #-----------
# # 轴指向z+z- 的角度的pdf
# #  path2 = 'D:\\guan2019\\2_ball\\2_data\\60Hz_select_rot\\'
# fig = plt.figure()
# ax1 = plt.subplot(111)
# label = []
# for i in range(len(filename)):  # 2,3):    #
#     # plt.scatter(pdf_rot_dict[filename[i] + 'x_theta'], pdf_rot_dict[filename[i] + 'p_theta'], alpha=0.5,
#     #             color=colors[i], cmap='hsv', label=filename[i] + 'g')
#     plt.plot(pdf_rot_dict[filename[i] + 'x_theta'], pdf_rot_dict[filename[i] + 'p_theta'], 'o-',alpha=0.5,
#                 color=colors[i],  label=filename[i] + 'g')
#     label.append(filename[i] + 'g')
# plt.legend(label)
# ax1.set_title('PDF -- the Value of Angular Displacement' + ' [60Hz] ', fontsize=10)  # the Value of
# ax1.set_xlabel('$\Delta \Theta(°)$')
# ax1.set_ylabel('P')
# ax1.plot()
# plt.show()
# del fig, ax1

#-----------
# 轴指向z+ 的角度的pdf
#  path2 = 'D:\\guan2019\\2_ball\\2_data\\60Hz_select_rot\\'
fig = plt.figure()
ax1 = plt.subplot(111)
label = []
for i in range(len(filename)):  # 2,3):    #
    plt.plot(pdf_rot_dict[filename[i] + 'x_new_theta'], pdf_rot_dict[filename[i] + 'p_new_theta'], alpha=0.5,color=colors[i])
    plt.scatter(pdf_rot_dict[filename[i] + 'x_new_theta'], pdf_rot_dict[filename[i] + 'p_new_theta'], marker=marker[i], c='', s=25, edgecolor=colors[i], label=filename[i] + 'g')
    ax1.errorbar(pdf_rot_dict[filename[i] + 'x_new_theta'], pdf_rot_dict[filename[i] + 'p_new_theta'],yerr=err_rot_dict[filename[i]],fmt='none',elinewidth=1,ms=1,ecolor=colors[i])
    label.append(filename[i] + 'g')
leg1 = plt.legend(label)
leg1.get_frame().set_linewidth(0.0)
ax1.xaxis.set_minor_locator(MultipleLocator(1))
ax1.yaxis.set_minor_locator(MultipleLocator(0.01))
ax1.xaxis.set_major_locator(MultipleLocator(5))
ax1.tick_params(axis="x", direction="in")
ax1.tick_params(axis="y", direction="in")
ax1.tick_params(which='minor', direction='in')
ax1.set_title('PDF -- Angular Displacement' + ' [60Hz] ', fontsize=10)  # the Value of
ax1.set_xlabel('$\Delta \Theta(°)$')
ax1.set_ylabel('P')
ax1.plot()
plt.yscale('log')
plt.show()
del fig, ax1

# #-----------
# # energy (轴指向z+ 的角度)的pdf 没写
fig = plt.figure()
ax1 = plt.subplot(111)
label = []
for i in range(len(filename)):  # 2,3):    #
    plt.plot(pdf_rot_dict[filename[i] + 'x_rot_energy'], pdf_rot_dict[filename[i] + 'p_rot_energy'], alpha=0.5,color=colors[i]) 
    plt.scatter(pdf_rot_dict[filename[i] + 'x_rot_energy'], pdf_rot_dict[filename[i] + 'p_rot_energy'], marker=marker[i], c='', s=25, edgecolor=colors[i], label=filename[i] + 'g')
    ax1.errorbar(pdf_rot_dict[filename[i] + 'x_rot_energy'], pdf_rot_dict[filename[i] + 'p_rot_energy'],yerr=err_rotenergy_dict[filename[i]][-len(pdf_rot_dict[filename[i] + 'p_rot_energy']):],fmt='none',elinewidth=1,ms=1,ecolor=colors[i])
    label.append(filename[i] + 'g')
leg1 = plt.legend(label)
leg1.get_frame().set_linewidth(0.0)
ax1.xaxis.set_minor_locator(MultipleLocator(1e-9))
# ax1.yaxis.set_minor_locator(MultipleLocator(0.01))
# ax1.xaxis.set_major_locator(MultipleLocator(5))
ax1.tick_params(axis="x", direction="in")
ax1.tick_params(axis="y", direction="in")
ax1.tick_params(which='minor', direction='in')
ax1.set_title('PDF -- Angular Energy Displacement' + ' [60Hz] ', fontsize=10)  # the Value of
ax1.set_xlabel('$E_R (J)$')
ax1.set_ylabel('P')
ax1.plot()
plt.yscale('log')
plt.show()
del fig, ax1


#-----------
# euler1的pdf
#  path2 = 'D:\\guan2019\\2_ball\\2_data\\60Hz_select_rot\\'
fig = plt.figure()
ax1 = plt.subplot(111)
label = []
for i in range(len(filename)):  # 2,3):    #
    # plt.scatter(pdf_rot_dict[filename[i] + 'x_theta'], pdf_rot_dict[filename[i] + 'p_theta'], alpha=0.5,
    #             color=colors[i], cmap='hsv', label=filename[i] + 'g')
    plt.plot(pdf_rot_dict[filename[i] + 'x_euler1'], pdf_rot_dict[filename[i] + 'p_euler1'], 'o-',alpha=0.5,
                color=colors[i],  label=filename[i] + 'g')
    label.append(filename[i] + 'g')
plt.legend(label)
ax1.set_title(r'PDF -- Euler $\alpha$' + ' [60Hz] ', fontsize=10)  # the Value of
ax1.set_xlabel(r'$\Delta \alpha(°)$')
ax1.set_ylabel('P')
ax1.plot()
plt.yscale('log')
plt.show()
del fig, ax1

# euler2的pdf
#  path2 = 'D:\\guan2019\\2_ball\\2_data\\60Hz_select_rot\\'
fig = plt.figure()
ax1 = plt.subplot(111)
label = []
for i in range(len(filename)):  # 2,3):    #
    # plt.scatter(pdf_rot_dict[filename[i] + 'x_theta'], pdf_rot_dict[filename[i] + 'p_theta'], alpha=0.5,
    #             color=colors[i], cmap='hsv', label=filename[i] + 'g')
    plt.plot(pdf_rot_dict[filename[i] + 'x_euler2'], pdf_rot_dict[filename[i] + 'p_euler2'], 'o-',alpha=0.5,
                color=colors[i],  label=filename[i] + 'g')
    label.append(filename[i] + 'g')
plt.legend(label)
ax1.set_title(r'PDF -- Euler $\beta $' + ' [60Hz] ', fontsize=10)  # the Value of
ax1.set_xlabel(r'$\Delta \beta (°)$')
ax1.set_ylabel('P')
ax1.plot()
plt.yscale('log')
plt.show()
del fig, ax1

# euler3的pdf
#  path2 = 'D:\\guan2019\\2_ball\\2_data\\60Hz_select_rot\\'
fig = plt.figure()
ax1 = plt.subplot(111)
label = []
for i in range(len(filename)):  # 2,3):    #
    # plt.scatter(pdf_rot_dict[filename[i] + 'x_theta'], pdf_rot_dict[filename[i] + 'p_theta'], alpha=0.5,
    #             color=colors[i], cmap='hsv', label=filename[i] + 'g')
    plt.plot(pdf_rot_dict[filename[i] + 'x_euler3'], pdf_rot_dict[filename[i] + 'p_euler3'], 'o-',alpha=0.5,
                color=colors[i],  label=filename[i] + 'g')
    label.append(filename[i] + 'g')
plt.legend(label)
ax1.set_title(r'PDF -- Euler $\gamma$' + ' [60Hz] ', fontsize=10)  # the Value of
ax1.set_xlabel(r'$\Delta \gamma(°)$')
ax1.set_ylabel('P')
ax1.plot()
plt.yscale('log')
plt.show()
del fig, ax1

# # #-----------
# # quaternion1的pdf
# #  path2 = 'D:\\guan2019\\2_ball\\2_data\\60Hz_select_rot\\'
# fig = plt.figure()
# ax1 = plt.subplot(111)
# label = []
# for i in range(len(filename)):  # 2,3):    #
#     # plt.scatter(pdf_rot_dict[filename[i] + 'x_theta'], pdf_rot_dict[filename[i] + 'p_theta'], alpha=0.5,
#     #             color=colors[i], cmap='hsv', label=filename[i] + 'g')
#     plt.plot(pdf_rot_dict[filename[i] + 'x_quat1'], pdf_rot_dict[filename[i] + 'p_quat1'], 'o-',alpha=0.5,
#                 color=colors[i],  label=filename[i] + 'g')
#     label.append(filename[i] + 'g')
# plt.legend(label)
# ax1.set_title(r'PDF -- quaternion1' + ' [60Hz] ', fontsize=10)  # the Value of
# ax1.set_xlabel(r'quaternion1')
# ax1.set_ylabel('P')
# ax1.plot()
# # plt.yscale('log')
# plt.show()
# del fig, ax1

# # quaternion2的pdf
# fig = plt.figure()
# ax1 = plt.subplot(111)
# label = []
# for i in range(len(filename)):  # 2,3):    #
#     # plt.scatter(pdf_rot_dict[filename[i] + 'x_theta'], pdf_rot_dict[filename[i] + 'p_theta'], alpha=0.5,
#     #             color=colors[i], cmap='hsv', label=filename[i] + 'g')
#     plt.plot(pdf_rot_dict[filename[i] + 'x_quat2'], pdf_rot_dict[filename[i] + 'p_quat2'], 'o-',alpha=0.5,
#                 color=colors[i],  label=filename[i] + 'g')
#     label.append(filename[i] + 'g')
# plt.legend(label)
# ax1.set_title(r'PDF -- quaternion2' + ' [60Hz] ', fontsize=10)  # the Value of
# ax1.set_xlabel(r'quaternion2')
# ax1.set_ylabel('P')
# # plt.yscale('log')
# ax1.plot()
# plt.show()
# del fig, ax1

# # quaternion2的pdf
# fig = plt.figure()
# ax1 = plt.subplot(111)
# label = []
# for i in range(len(filename)):  # 2,3):    #
#     # plt.scatter(pdf_rot_dict[filename[i] + 'x_theta'], pdf_rot_dict[filename[i] + 'p_theta'], alpha=0.5,
#     #             color=colors[i], cmap='hsv', label=filename[i] + 'g')
#     plt.plot(pdf_rot_dict[filename[i] + 'x_quat3'], pdf_rot_dict[filename[i] + 'p_quat3'], 'o-',alpha=0.5,
#                 color=colors[i],  label=filename[i] + 'g')
#     label.append(filename[i] + 'g')
# plt.legend(label)
# ax1.set_title(r'PDF -- quaternion3' + ' [60Hz] ', fontsize=10)  # the Value of
# ax1.set_xlabel(r'quaternion3')
# ax1.set_ylabel('P')
# ax1.plot()
# # plt.yscale('log')
# plt.show()
# del fig, ax1

# # quaternion4的pdf
# fig = plt.figure()
# ax1 = plt.subplot(111)
# label = []
# for i in range(len(filename)):  # 2,3):    #
#     # plt.scatter(pdf_rot_dict[filename[i] + 'x_theta'], pdf_rot_dict[filename[i] + 'p_theta'], alpha=0.5,
#     #             color=colors[i], cmap='hsv', label=filename[i] + 'g')
#     plt.plot(pdf_rot_dict[filename[i] + 'x_quat4'], pdf_rot_dict[filename[i] + 'p_quat4'], 'o-',alpha=0.5,
#                 color=colors[i],  label=filename[i] + 'g')
#     label.append(filename[i] + 'g')
# plt.legend(label)
# ax1.set_title(r'PDF -- quaternion4' + ' [60Hz] ', fontsize=10)  # the Value of
# ax1.set_xlabel(r'quaternion4')
# ax1.set_ylabel('P')
# ax1.plot()
# # plt.yscale('log')
# plt.show()
# del fig, ax1